from __future__ import annotations

import csv
import io
from pathlib import Path
from typing import Any, Iterable


MAX_SPREADSHEET_BYTES = 5 * 1024 * 1024
MAX_SPREADSHEET_ROWS = 2000
MAX_PARSED_PROMPTS = 200
MAX_PROMPT_CHARACTERS = 4000
SUPPORTED_SPREADSHEET_SUFFIXES = {".xlsx", ".xls", ".ods", ".csv", ".tsv", ".txt"}
PROMPT_HEADERS = {
    "视频提示词",
    "提示词",
    "视频文案",
    "生成提示词",
    "prompt",
    "video prompt",
    "video_prompt",
    "description",
    "文案",
    "内容",
}


class SpreadsheetImportError(ValueError):
    pass


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return " ".join(str(value).replace("\x00", "").split()).strip()


def _decode_text(data: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-16", "gb18030"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise SpreadsheetImportError("表格文本编码无法识别，请另存为 UTF-8 CSV 后重试")


def _csv_rows(data: bytes, suffix: str) -> list[list[Any]]:
    text = _decode_text(data)
    if suffix == ".txt":
        return [[line] for line in text.splitlines()]
    delimiter = "\t" if suffix == ".tsv" else ","
    try:
        dialect = csv.Sniffer().sniff(text[:8192], delimiters=",\t;|")
        delimiter = dialect.delimiter
    except csv.Error:
        pass
    return [list(row) for row in csv.reader(io.StringIO(text), delimiter=delimiter)]


def _xlsx_rows(data: bytes) -> list[list[Any]]:
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        sheet = workbook.active
        return [list(row) for row in sheet.iter_rows(max_row=MAX_SPREADSHEET_ROWS, max_col=80, values_only=True)]
    except Exception as exc:
        raise SpreadsheetImportError("XLSX 文件无法读取或文件已损坏") from exc


def _xls_rows(data: bytes) -> list[list[Any]]:
    try:
        import xlrd

        workbook = xlrd.open_workbook(file_contents=data, on_demand=True)
        sheet = workbook.sheet_by_index(0)
        return [sheet.row_values(index, 0, min(sheet.ncols, 80)) for index in range(min(sheet.nrows, MAX_SPREADSHEET_ROWS))]
    except Exception as exc:
        raise SpreadsheetImportError("XLS 文件无法读取或文件已损坏") from exc


def _ods_rows(data: bytes) -> list[list[Any]]:
    try:
        from odf import teletype
        from odf.opendocument import load
        from odf.table import Table, TableCell, TableRow

        document = load(io.BytesIO(data))
        sheets = document.spreadsheet.getElementsByType(Table)
        if not sheets:
            return []
        rows: list[list[Any]] = []
        for row in sheets[0].getElementsByType(TableRow)[:MAX_SPREADSHEET_ROWS]:
            values = []
            for cell in row.getElementsByType(TableCell)[:80]:
                values.append(teletype.extractText(cell))
            rows.append(values)
        return rows
    except Exception as exc:
        raise SpreadsheetImportError("ODS 文件无法读取或文件已损坏") from exc


def _prompt_column(rows: list[list[Any]]) -> tuple[int, int]:
    for row_index, row in enumerate(rows[:10]):
        for column_index, value in enumerate(row[:80]):
            if _text(value).lower() in PROMPT_HEADERS:
                return row_index + 1, column_index
    max_columns = min(80, max((len(row) for row in rows), default=0))
    if not max_columns:
        return 0, 0
    scores: list[tuple[float, int]] = []
    for column_index in range(max_columns):
        values = [_text(row[column_index]) for row in rows[:MAX_SPREADSHEET_ROWS] if column_index < len(row)]
        values = [value for value in values if value and value.lower() not in PROMPT_HEADERS]
        text_values = [value for value in values if not value.replace(".", "", 1).isdigit()]
        average_length = sum(map(len, text_values)) / max(1, len(text_values))
        scores.append((len(text_values) * 10 + min(average_length, 200), column_index))
    return 0, max(scores)[1]


def extract_prompts(rows: Iterable[Iterable[Any]]) -> list[dict[str, Any]]:
    normalized_rows = [list(row)[:80] for row in rows][:MAX_SPREADSHEET_ROWS]
    start_row, column_index = _prompt_column(normalized_rows)
    prompts: list[dict[str, Any]] = []
    seen: set[str] = set()
    for row_index, row in enumerate(normalized_rows[start_row:], start=start_row + 1):
        if column_index >= len(row):
            continue
        prompt = _text(row[column_index])
        if not prompt or prompt.lower() in PROMPT_HEADERS:
            continue
        prompt = prompt[:MAX_PROMPT_CHARACTERS]
        fingerprint = prompt.casefold()
        if fingerprint in seen:
            continue
        seen.add(fingerprint)
        prompts.append({"row": row_index, "prompt": prompt})
        if len(prompts) >= MAX_PARSED_PROMPTS:
            break
    if not prompts:
        raise SpreadsheetImportError("未在表格中识别到视频提示词")
    return prompts


def parse_spreadsheet(filename: str, data: bytes) -> list[dict[str, Any]]:
    suffix = Path(filename or "").suffix.lower()
    if suffix not in SUPPORTED_SPREADSHEET_SUFFIXES:
        raise SpreadsheetImportError("不支持此表格格式，请使用 XLSX、XLS、ODS、CSV、TSV 或 TXT")
    if not data:
        raise SpreadsheetImportError("导入的表格为空")
    if len(data) > MAX_SPREADSHEET_BYTES:
        raise SpreadsheetImportError("表格文件不能超过 5MB")
    if suffix == ".xlsx":
        rows = _xlsx_rows(data)
    elif suffix == ".xls":
        rows = _xls_rows(data)
    elif suffix == ".ods":
        rows = _ods_rows(data)
    else:
        rows = _csv_rows(data, suffix)
    return extract_prompts(rows)
